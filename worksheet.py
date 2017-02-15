import openpyxl

class PyXlWorksheet(object):

    def __init__(self, worksheet):
        self.headers = dict()
        for col_num in range(0, worksheet.max_column):
            header = worksheet.cell(row=1, column=col_num + 1).value
            self.headers.update({ header : col_num + 1 })

        self.data = list()
        for row_num in range(0, worksheet.max_row + 1):
            row_list = list()
            for col_num in range(0, worksheet.max_column + 1):
                row_list.append(worksheet.cell(row=row_num, column=col_num).value)
            self.data.append(row_list)

    def row(self, row_num):
        row_dict = dict()
        for col in self.data[row_num]:
            row_dict.update()

